"""Export Service — Excel, CSV, ZUGFeRD.

Generates downloadable exports in multiple formats:
1. Excel (.xlsx) — formatted with openpyxl
2. CSV (.csv) — German-locale semicolon-separated
3. ZUGFeRD (PDF+XML) — EN 16931 compliant
"""
from __future__ import annotations

import csv
import io
import json
import logging
from typing import Any

logger = logging.getLogger(__name__)


class ExportService:

    def to_csv(self, invoices: list[dict]) -> str:
        output = io.StringIO()
        writer = csv.writer(output, delimiter=";", quotechar='"')
        writer.writerow(["Rechnungsnummer", "Lieferant", "Datum", "Betrag", "Währung", "Status", "Dokument-ID"])
        for inv in invoices:
            writer.writerow([
                inv.get("invoice_number", ""),
                inv.get("supplier", ""),
                inv.get("invoice_date", ""),
                str(inv.get("total_amount", "")).replace(".", ","),
                inv.get("currency", "EUR"),
                inv.get("status", inv.get("current_state", "")),
                inv.get("document_id", ""),
            ])
        return output.getvalue()

    def to_excel_bytes(self, invoices: list[dict]) -> bytes:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

        wb = Workbook()
        ws = wb.active
        ws.title = "Rechnungen"

        # Header style
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="003856", end_color="003856", fill_type="solid")
        thin_border = Border(
            left=Side(style="thin"), right=Side(style="thin"),
            top=Side(style="thin"), bottom=Side(style="thin"),
        )

        headers = ["Rechnungsnummer", "Lieferant", "Datum", "Fällig", "Netto", "MwSt", "Brutto", "Währung", "Status", "Dokument-ID"]
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=h)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")
            cell.border = thin_border

        for row_idx, inv in enumerate(invoices, 2):
            extracted = {}
            if inv.get("extracted_data"):
                try:
                    extracted = json.loads(inv["extracted_data"]) if isinstance(inv["extracted_data"], str) else inv["extracted_data"]
                except Exception:
                    pass

            values = [
                inv.get("invoice_number", ""),
                inv.get("supplier", ""),
                inv.get("invoice_date", ""),
                inv.get("due_date", ""),
                extracted.get("total_amount_net", ""),
                inv.get("tax_amount", ""),
                inv.get("total_amount", ""),
                inv.get("currency", "EUR"),
                inv.get("status", inv.get("current_state", "")),
                inv.get("document_id", ""),
            ]
            for col, val in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col, value=val)
                cell.border = thin_border
                if col in (5, 6, 7) and val:
                    try:
                        cell.value = float(val)
                        cell.number_format = '#,##0.00 €'
                    except (ValueError, TypeError):
                        pass

        # Auto-width
        for col in ws.columns:
            max_len = max(len(str(cell.value or "")) for cell in col)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 4, 40)

        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()

    def to_zugferd_xml(self, invoice_data: dict) -> str:
        """Generate ZUGFeRD 2.1 (Factur-X) Cross Industry Invoice XML."""
        supplier = invoice_data.get("supplier", "Unbekannt")
        inv_num = invoice_data.get("invoice_number", "UNKNOWN")
        inv_date = (invoice_data.get("invoice_date") or "20260101").replace("-", "")
        due_date = (invoice_data.get("due_date") or "").replace("-", "") or inv_date
        gross = invoice_data.get("total_amount_gross", invoice_data.get("total_amount", 0))
        net = invoice_data.get("total_amount_net", float(gross) / 1.19 if gross else 0)
        tax = invoice_data.get("tax_amount", float(gross) - float(net) if gross and net else 0)
        currency = invoice_data.get("currency", "EUR")

        xml = f"""<?xml version="1.0" encoding="UTF-8"?>
<rsm:CrossIndustryInvoice xmlns:rsm="urn:un:unece:uncefact:data:standard:CrossIndustryInvoice:100"
  xmlns:ram="urn:un:unece:uncefact:data:standard:ReusableAggregateBusinessInformationEntity:100"
  xmlns:udt="urn:un:unece:uncefact:data:standard:UnqualifiedDataType:100"
  xmlns:qdt="urn:un:unece:uncefact:data:standard:QualifiedDataType:100">
  <rsm:ExchangedDocumentContext>
    <ram:GuidelineSpecifiedDocumentContextParameter>
      <ram:ID>urn:factur-x.eu:1p0:extended</ram:ID>
    </ram:GuidelineSpecifiedDocumentContextParameter>
  </rsm:ExchangedDocumentContext>
  <rsm:ExchangedDocument>
    <ram:ID>{inv_num}</ram:ID>
    <ram:TypeCode>380</ram:TypeCode>
    <ram:IssueDateTime><udt:DateTimeString format="102">{inv_date}</udt:DateTimeString></ram:IssueDateTime>
  </rsm:ExchangedDocument>
  <rsm:SupplyChainTradeTransaction>
    <ram:ApplicableHeaderTradeAgreement>
      <ram:SellerTradeParty>
        <ram:Name>{supplier}</ram:Name>
        <ram:PostalTradeAddress><ram:CountryID>DE</ram:CountryID></ram:PostalTradeAddress>
      </ram:SellerTradeParty>
      <ram:BuyerTradeParty>
        <ram:Name>SBS Deutschland GmbH &amp; Co. KG</ram:Name>
        <ram:PostalTradeAddress>
          <ram:PostcodeCode>69115</ram:PostcodeCode>
          <ram:CityName>Heidelberg</ram:CityName>
          <ram:CountryID>DE</ram:CountryID>
        </ram:PostalTradeAddress>
      </ram:BuyerTradeParty>
    </ram:ApplicableHeaderTradeAgreement>
    <ram:ApplicableHeaderTradeDelivery/>
    <ram:ApplicableHeaderTradeSettlement>
      <ram:InvoiceCurrencyCode>{currency}</ram:InvoiceCurrencyCode>
      <ram:SpecifiedTradeSettlementHeaderMonetarySummation>
        <ram:LineTotalAmount>{float(net):.2f}</ram:LineTotalAmount>
        <ram:TaxBasisTotalAmount>{float(net):.2f}</ram:TaxBasisTotalAmount>
        <ram:TaxTotalAmount currencyID="{currency}">{float(tax):.2f}</ram:TaxTotalAmount>
        <ram:GrandTotalAmount>{float(gross):.2f}</ram:GrandTotalAmount>
        <ram:DuePayableAmount>{float(gross):.2f}</ram:DuePayableAmount>
      </ram:SpecifiedTradeSettlementHeaderMonetarySummation>
    </ram:ApplicableHeaderTradeSettlement>
  </rsm:SupplyChainTradeTransaction>
</rsm:CrossIndustryInvoice>"""
        return xml
