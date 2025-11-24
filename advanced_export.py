"""
Advanced Export Functionality
Erweiterte Excel-Exports mit allen Features
"""
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils.dataframe import dataframe_to_rows
import sqlite3
from datetime import datetime
from typing import List, Dict
import io

def create_comprehensive_excel(job_id: str) -> bytes:
    """
    Erstellt umfassendes Excel mit mehreren Sheets
    """
    conn = sqlite3.connect('invoices.db', check_same_thread=False)
    
    # Sheet 1: Alle Rechnungen
    invoices_df = pd.read_sql_query('''
        SELECT 
            rechnungsnummer as "Rechnungsnummer",
            datum as "Datum",
            rechnungsaussteller as "Aussteller",
            rechnungsempfaenger as "Empfänger",
            betrag_netto as "Netto",
            mwst_betrag as "MwSt",
            betrag_brutto as "Brutto",
            waehrung as "Währung",
            detected_language as "Sprache"
        FROM invoices
        WHERE job_id = ?
        ORDER BY datum DESC
    ''', conn, params=(job_id,))
    
    # Sheet 2: Duplikate (falls vorhanden)
    try:
        duplicates_df = pd.read_sql_query('''
            SELECT 
                i1.rechnungsnummer as "Rechnung 1",
                i2.rechnungsnummer as "Rechnung 2",
                dd.match_reason as "Grund",
                dd.status as "Status"
            FROM duplicate_detections dd
            JOIN invoices i1 ON dd.invoice_id_1 = i1.id
            JOIN invoices i2 ON dd.invoice_id_2 = i2.id
            WHERE dd.job_id = ?
        ''', conn, params=(job_id,))
    except:
        duplicates_df = pd.DataFrame()
    
    # Sheet 3: Plausibilitäts-Warnungen (falls vorhanden)
    try:
        plausibility_df = pd.read_sql_query('''
            SELECT 
                i.rechnungsnummer as "Rechnung",
                pc.check_type as "Check-Typ",
                pc.severity as "Schwere",
                pc.details as "Details",
                pc.status as "Status"
            FROM plausibility_checks pc
            JOIN invoices i ON pc.invoice_id = i.id
            WHERE pc.job_id = ?
            ORDER BY pc.severity DESC
        ''', conn, params=(job_id,))
    except:
        plausibility_df = pd.DataFrame()
    
    # Sheet 4: Aussteller-Statistiken
    issuer_stats_df = pd.read_sql_query('''
        SELECT 
            rechnungsaussteller as "Aussteller",
            COUNT(*) as "Anzahl",
            SUM(betrag_brutto) as "Gesamt Brutto",
            AVG(betrag_brutto) as "Ø Brutto",
            MIN(datum) as "Erste Rechnung",
            MAX(datum) as "Letzte Rechnung"
        FROM invoices
        WHERE job_id = ?
        GROUP BY rechnungsaussteller
        ORDER BY COUNT(*) DESC
    ''', conn, params=(job_id,))
    
    conn.close()
    
    # Erstelle Excel
    output = io.BytesIO()
    
    with pd.ExcelWriter(output, engine='openpyxl') as writer:
        # Sheet 1: Rechnungen
        invoices_df.to_excel(writer, sheet_name='Rechnungen', index=False)
        
        # Sheet 2: Duplikate (nur wenn vorhanden)
        if not duplicates_df.empty:
            duplicates_df.to_excel(writer, sheet_name='Duplikate', index=False)
        
        # Sheet 3: Plausibility (nur wenn vorhanden)
        if not plausibility_df.empty:
            plausibility_df.to_excel(writer, sheet_name='Warnungen', index=False)
        
        # Sheet 4: Statistiken
        issuer_stats_df.to_excel(writer, sheet_name='Aussteller-Stats', index=False)
        
        # Formatierung
        workbook = writer.book
        ws = workbook['Rechnungen']
        
        # Header-Style
        header_fill = PatternFill(start_color='003856', end_color='003856', fill_type='solid')
        header_font = Font(color='FFFFFF', bold=True)
        
        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal='center')
        
        # Auto-width
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
    
    output.seek(0)
    return output.read()

def create_zip_export(job_id: str) -> bytes:
    """
    Erstellt ZIP mit Excel + JSON
    """
    import zipfile
    import json
    from pathlib import Path
    
    output = io.BytesIO()
    
    with zipfile.ZipFile(output, 'w', zipfile.ZIP_DEFLATED) as zipf:
        # Excel hinzufügen
        excel_bytes = create_comprehensive_excel(job_id)
        zipf.writestr(f'report_{job_id[:8]}.xlsx', excel_bytes)
        
        # JSON-Export
        conn = sqlite3.connect('invoices.db', check_same_thread=False)
        conn.row_factory = sqlite3.Row
        cursor = conn.cursor()
        
        cursor.execute('SELECT * FROM invoices WHERE job_id = ?', (job_id,))
        invoices = [dict(row) for row in cursor.fetchall()]
        
        json_data = json.dumps(invoices, indent=2, default=str)
        zipf.writestr(f'invoices_{job_id[:8]}.json', json_data)
        
        # PDFs hinzufügen (falls vorhanden)
        pdf_dir = Path(f'uploads/{job_id}')
        if pdf_dir.exists():
            for pdf_file in pdf_dir.glob('*.pdf'):
                zipf.write(pdf_file, f'pdfs/{pdf_file.name}')
        
        conn.close()
    
    output.seek(0)
    return output.read()
