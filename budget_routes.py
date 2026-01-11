"""
Budget API Routes für SBS Invoice System
"""
from fastapi import APIRouter, Request, Query, HTTPException
from fastapi.responses import HTMLResponse, FileResponse
from fastapi.templating import Jinja2Templates
from pydantic import BaseModel
from typing import Optional
from datetime import datetime

from budget_service import BudgetService

router = APIRouter(tags=["Budget"])
templates = Jinja2Templates(directory="web/templates")
budget_service = BudgetService()

MONAT_NAMEN = ["Januar", "Februar", "März", "April", "Mai", "Juni", "Juli", "August", "September", "Oktober", "November", "Dezember"]

class BudgetSetRequest(BaseModel):
    kategorie_id: int
    jahr: int
    monat: int
    betrag: float
    notiz: Optional[str] = None

class BudgetCopyRequest(BaseModel):
    von_jahr: int
    von_monat: int
    nach_jahr: int
    nach_monat: int
    prozent_aenderung: float = 0

@router.get("/budget", response_class=HTMLResponse)
async def budget_dashboard(request: Request, jahr: int = Query(default=None), monat: int = Query(default=None)):
    if jahr is None:
        jahr = datetime.now().year
    if monat is None:
        monat = datetime.now().month
    
    budget_service.update_ist_werte_from_invoices(jahr, monat)
    summary = budget_service.get_dashboard_summary(jahr, monat)
    trend_data = budget_service.get_trend_data(monate_zurueck=12)
    pie_data = budget_service.get_kategorie_verteilung(jahr, monat)
    
    # Integration: Bezahlt/Offen Status laden
    try:
        conn = budget_service._get_connection()
        cursor = conn.cursor()
        cursor.execute("""
            SELECT 
                COALESCE(SUM(CASE WHEN payment_status = 'paid' THEN betrag_brutto ELSE 0 END), 0) as bezahlt,
                COALESCE(SUM(CASE WHEN payment_status != 'paid' OR payment_status IS NULL THEN betrag_brutto ELSE 0 END), 0) as offen,
                COUNT(CASE WHEN payment_status = 'paid' THEN 1 END) as anz_bezahlt,
                COUNT(CASE WHEN payment_status != 'paid' OR payment_status IS NULL THEN 1 END) as anz_offen
            FROM invoices 
            WHERE datum IS NOT NULL 
                AND strftime('%Y', datum) = ? 
                AND strftime('%m', datum) = ?
        """, (str(jahr), str(monat).zfill(2)))
        row = cursor.fetchone()
        summary["gesamt_bezahlt"] = round(row["bezahlt"], 2) if row else 0
        summary["gesamt_offen"] = round(row["offen"], 2) if row else 0
        summary["anzahl_bezahlt"] = row["anz_bezahlt"] if row else 0
        summary["anzahl_offen"] = row["anz_offen"] if row else 0
        conn.close()
    except Exception as e:
        summary["gesamt_bezahlt"] = 0
        summary["gesamt_offen"] = 0
        summary["anzahl_bezahlt"] = 0
        summary["anzahl_offen"] = 0
    
    return templates.TemplateResponse("budget.html", {
        "request": request, "jahr": jahr, "monat": monat,
        "monat_namen": MONAT_NAMEN, "summary": summary,
        "trend_data": trend_data, "pie_data": pie_data
    })


@router.get("/budget/jahr", response_class=HTMLResponse)
async def budget_jahresansicht(request: Request, jahr: int = Query(default=None)):
    if jahr is None:
        jahr = datetime.now().year
    jahresbudget = budget_service.get_jahresbudget(jahr)
    return templates.TemplateResponse("budget_jahr.html", {
        "request": request, "jahr": jahr, "monat_namen": MONAT_NAMEN, "jahresbudget": jahresbudget
    })

@router.get("/api/budget/kategorien")
async def get_kategorien(nur_aktive: bool = Query(default=True), mit_budget: bool = Query(default=False), jahr: int = Query(default=None), monat: int = Query(default=None)):
    kategorien = budget_service.get_kategorien(nur_aktive)
    if mit_budget and jahr and monat:
        auswertungen = budget_service.get_monatsauswertung(jahr, monat)
        budget_map = {a["kategorie_id"]: a["budget_soll"] for a in auswertungen}
        for kat in kategorien:
            kat["budget"] = budget_map.get(kat["id"], 0)
    return kategorien

@router.post("/api/budget/set")
async def set_budget(request: BudgetSetRequest):
    try:
        budget_id = budget_service.set_monatsbudget(kategorie_id=request.kategorie_id, jahr=request.jahr, monat=request.monat, betrag=request.betrag, notiz=request.notiz)
        return {"success": True, "budget_id": budget_id}
    except Exception as e:
        return {"success": False, "error": str(e)}

@router.post("/api/budget/copy")
async def copy_budgets(request: BudgetCopyRequest):
    try:
        kopiert = budget_service.copy_budgets_to_month(von_jahr=request.von_jahr, von_monat=request.von_monat, nach_jahr=request.nach_jahr, nach_monat=request.nach_monat, prozent_aenderung=request.prozent_aenderung)
        return {"success": True, "kopiert": kopiert}
    except Exception as e:
        return {"success": False, "error": str(e)}

@router.get("/api/budget/summary")
async def get_summary(jahr: int = Query(...), monat: int = Query(...)):
    budget_service.update_ist_werte_from_invoices(jahr, monat)
    return budget_service.get_dashboard_summary(jahr, monat)

@router.get("/api/budget/trend")
async def get_trend(kategorie_id: Optional[int] = Query(default=None), monate: int = Query(default=12)):
    return budget_service.get_trend_data(kategorie_id, monate)

@router.get("/api/budget/verteilung")
async def get_verteilung(jahr: int = Query(...), monat: int = Query(...)):
    return budget_service.get_kategorie_verteilung(jahr, monat)

@router.get("/api/budget/alerts")
async def get_alerts(nur_ungelesen: bool = Query(default=False), limit: int = Query(default=50)):
    return budget_service.get_alerts(nur_ungelesen, limit)

@router.post("/api/budget/alerts/{alert_id}/read")
async def mark_alert_read(alert_id: int):
    return {"success": budget_service.mark_alert_read(alert_id)}

@router.post("/api/budget/alerts/read-all")
async def mark_all_alerts_read():
    count = budget_service.mark_all_alerts_read()
    return {"success": True, "count": count}

@router.get("/api/budget/export")
async def export_budgets(jahr: int = Query(...), monat: int = Query(...)):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Border, Side
        
        auswertungen = budget_service.get_monatsauswertung(jahr, monat)
        summary = budget_service.get_dashboard_summary(jahr, monat)
        
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = f"Budget {MONAT_NAMEN[monat-1]} {jahr}"
        
        header_font = Font(bold=True, color="FFFFFF")
        header_fill = PatternFill(start_color="003856", end_color="003856", fill_type="solid")
        thin_border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
        
        ws.merge_cells('A1:F1')
        ws['A1'] = f"Budget-Report: {MONAT_NAMEN[monat-1]} {jahr}"
        ws['A1'].font = Font(bold=True, size=16)
        
        ws['A3'], ws['B3'] = "Gesamt-Budget:", summary['gesamt_budget']
        ws['A4'], ws['B4'] = "Gesamt-Ausgaben:", summary['gesamt_ausgaben']
        ws['B3'].number_format = ws['B4'].number_format = '#,##0.00 €'
        
        headers = ['Kategorie', 'Budget', 'Ist', 'Differenz', 'Auslastung', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=6, column=col, value=header)
            cell.font, cell.fill, cell.border = header_font, header_fill, thin_border
        
        for row, kat in enumerate(auswertungen, 7):
            ws.cell(row=row, column=1, value=kat['kategorie_name']).border = thin_border
            ws.cell(row=row, column=2, value=kat['budget_soll']).border = thin_border
            ws.cell(row=row, column=3, value=kat['ist_ausgaben']).border = thin_border
            ws.cell(row=row, column=4, value=kat['differenz']).border = thin_border
            ws.cell(row=row, column=5, value=f"{kat['auslastung_prozent']}%").border = thin_border
            status = "OK" if kat['alert_severity'] == 'info' else "Achtung" if kat['alert_severity'] == 'warning' else "Überschritten"
            ws.cell(row=row, column=6, value=status).border = thin_border
        
        ws.column_dimensions['A'].width = 28
        for c in 'BCDEF': ws.column_dimensions[c].width = 15
        
        filepath = f"/tmp/budget_{jahr}_{monat:02d}.xlsx"
        wb.save(filepath)
        return FileResponse(filepath, filename=f"budget_{jahr}_{monat:02d}.xlsx", media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
    except ImportError:
        raise HTTPException(status_code=500, detail="openpyxl nicht installiert")

@router.get("/api/budget/jahresbudget")
async def get_jahresbudget(jahr: int = Query(...)):
    return budget_service.get_jahresbudget(jahr)

@router.post("/api/budget/init-demo")
async def init_demo():
    from budget_service import create_demo_budgets
    create_demo_budgets()
    return {"success": True, "message": "Demo-Daten erstellt"}


# ==================== INTEGRATION MIT ZAHLUNGEN & KONTIERUNG ====================

@router.get("/api/budget/integration/uebersicht")
async def get_integration_uebersicht(jahr: int = Query(...), monat: int = Query(...)):
    """
    Kombinierte Übersicht: Budget + Zahlungsstatus + Kontierungen
    """
    from zahlungs_service import get_zahlungs_service
    
    # Budget-Daten
    budget_data = budget_service.get_monatsauswertung(jahr, monat)
    
    # Zahlungs-Daten aus DB
    conn = budget_service._get_connection()
    cursor = conn.cursor()
    
    # Bezahlte vs offene Beträge pro Kategorie
    cursor.execute("""
        SELECT 
            bk.id as kategorie_id,
            bk.name as kategorie_name,
            COALESCE(SUM(CASE WHEN i.payment_status = 'paid' THEN i.betrag_brutto ELSE 0 END), 0) as bezahlt,
            COALESCE(SUM(CASE WHEN i.payment_status != 'paid' OR i.payment_status IS NULL THEN i.betrag_brutto ELSE 0 END), 0) as offen,
            COUNT(CASE WHEN i.payment_status = 'paid' THEN 1 END) as anzahl_bezahlt,
            COUNT(CASE WHEN i.payment_status != 'paid' OR i.payment_status IS NULL THEN 1 END) as anzahl_offen
        FROM budget_kategorien bk
        LEFT JOIN kontierung_historie kh ON 1=1
        LEFT JOIN invoices i ON LOWER(i.rechnungsaussteller) LIKE '%' || LOWER(kh.lieferant_pattern) || '%'
            AND strftime('%Y', i.datum) = ?
            AND strftime('%m', i.datum) = ?
            AND (""" + " OR ".join([f"kh.final_account LIKE '{k}%'" for kat in budget_service.get_kategorien() for k in kat.get("konten_mapping", [])[:1]]) + """)
        WHERE bk.aktiv = 1
        GROUP BY bk.id
    """, (str(jahr), str(monat).zfill(2)))
    
    zahlungs_status = {row["kategorie_id"]: dict(row) for row in cursor.fetchall()}
    
    # Gesamt-Zahlungsstatus
    cursor.execute("""
        SELECT 
            COALESCE(SUM(CASE WHEN payment_status = 'paid' THEN betrag_brutto ELSE 0 END), 0) as gesamt_bezahlt,
            COALESCE(SUM(CASE WHEN payment_status != 'paid' THEN betrag_brutto ELSE 0 END), 0) as gesamt_offen,
            COUNT(CASE WHEN payment_status = 'paid' THEN 1 END) as anzahl_bezahlt,
            COUNT(CASE WHEN payment_status != 'paid' THEN 1 END) as anzahl_offen
        FROM invoices
        WHERE strftime('%Y', datum) = ? AND strftime('%m', datum) = ?
    """, (str(jahr), str(monat).zfill(2)))
    
    gesamt = cursor.fetchone()
    conn.close()
    
    # Kombiniere Budget mit Zahlungsstatus
    for kat in budget_data:
        zs = zahlungs_status.get(kat["kategorie_id"], {})
        kat["bezahlt"] = zs.get("bezahlt", 0)
        kat["offen"] = zs.get("offen", 0)
        kat["anzahl_bezahlt"] = zs.get("anzahl_bezahlt", 0)
        kat["anzahl_offen"] = zs.get("anzahl_offen", 0)
    
    return {
        "jahr": jahr,
        "monat": monat,
        "kategorien": budget_data,
        "gesamt_bezahlt": gesamt["gesamt_bezahlt"] if gesamt else 0,
        "gesamt_offen": gesamt["gesamt_offen"] if gesamt else 0,
        "anzahl_bezahlt": gesamt["anzahl_bezahlt"] if gesamt else 0,
        "anzahl_offen": gesamt["anzahl_offen"] if gesamt else 0
    }


@router.post("/api/budget/sync-from-kontierung")
async def sync_budget_from_kontierung():
    """
    Synchronisiert Budget-Ist-Werte aus allen Kontierungen
    """
    conn = budget_service._get_connection()
    cursor = conn.cursor()
    
    # Lösche alte Ist-Werte
    cursor.execute("DELETE FROM budget_ist_werte")
    
    # Hole alle Kategorien
    kategorien = budget_service.get_kategorien()
    
    synced = 0
    for kategorie in kategorien:
        konten = kategorie.get("konten_mapping", [])
        if not konten:
            continue
        
        konten_filter = " OR ".join([f"kh.final_account LIKE '{k}%'" for k in konten])
        
        cursor.execute(f"""
            SELECT 
                strftime('%Y', i.datum) as jahr,
                strftime('%m', i.datum) as monat,
                SUM(i.betrag_brutto) as summe,
                COUNT(*) as anzahl
            FROM invoices i
            INNER JOIN kontierung_historie kh 
                ON LOWER(i.rechnungsaussteller) LIKE '%' || LOWER(kh.lieferant_pattern) || '%'
            WHERE i.datum IS NOT NULL AND ({konten_filter})
            GROUP BY jahr, monat
        """)
        
        for row in cursor.fetchall():
            if row["jahr"] and row["monat"] and row["summe"]:
                cursor.execute("""
                    INSERT INTO budget_ist_werte (kategorie_id, jahr, monat, ist_betrag, anzahl_buchungen)
                    VALUES (?, ?, ?, ?, ?)
                """, (kategorie["id"], int(row["jahr"]), int(row["monat"]), row["summe"], row["anzahl"]))
                synced += 1
    
    conn.commit()
    conn.close()
    
    return {"success": True, "synced_entries": synced}


@router.get("/api/budget/offene-posten")
async def get_offene_posten(kategorie_id: int = Query(default=None)):
    """
    Zeigt offene (unbezahlte) Rechnungen pro Budget-Kategorie
    """
    conn = budget_service._get_connection()
    cursor = conn.cursor()
    
    query = """
        SELECT 
            i.id, i.rechnungsnummer, i.rechnungsaussteller, i.betrag_brutto,
            i.datum, i.faelligkeitsdatum, i.payment_status,
            bk.id as kategorie_id, bk.name as kategorie_name,
            kh.final_account
        FROM invoices i
        LEFT JOIN kontierung_historie kh 
            ON LOWER(i.rechnungsaussteller) LIKE '%' || LOWER(kh.lieferant_pattern) || '%'
        LEFT JOIN budget_kategorien bk ON 1=1
        WHERE i.payment_status != 'paid' OR i.payment_status IS NULL
    """
    
    if kategorie_id:
        kategorie = next((k for k in budget_service.get_kategorien() if k["id"] == kategorie_id), None)
        if kategorie:
            konten = kategorie.get("konten_mapping", [])
            if konten:
                konten_filter = " OR ".join([f"kh.final_account LIKE '{k}%'" for k in konten])
                query += f" AND ({konten_filter})"
    
    query += " ORDER BY i.faelligkeitsdatum ASC LIMIT 100"
    
    cursor.execute(query)
    posten = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return {"offene_posten": posten, "anzahl": len(posten)}
